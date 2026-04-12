"""
EFV (Federal Finance Administration) Debt Issuance Analysis
=============================================================
Downloads, parses, and saves structured data on Swiss Confederation
debt issuance across three instruments:

  1. GMBF (Geldmarktbuchforderungen / Money Market Debt Register Claims)
     - Short-term discount securities, maturities ~91d / ~182d / ~364d
     - Auctioned weekly on Tuesdays at 10:00-10:30 CET on SIX Repo Ltd platform
     - Settlement T+2 (Thursday value date)
     - Data: 2012–present

  2. Confederation Bonds (Eidgenössische Anleihen / Bundesanleihen)
     - Long-term fixed-rate bonds, maturities 4–50 years
     - Auctioned monthly (second Wednesday, August excluded) at 10:30-11:00 CET
     - Settlement approx. T+14 (two Fridays after auction)
     - Always two bonds (reopenings) per auction
     - Data: 2011–present

  3. Outstanding Bonds portfolio snapshot
     - Current face-value outstanding per ISIN including own-holdings breakdown
     - Updated by EFV periodically (typically end-of-quarter)

Data source: EFV (Eidgenössische Finanzverwaltung / Federal Finance Administration)
  https://www.efv.admin.ch/efv/en/home/finanzberichterstattung/

Author: SNB_MoPo project
"""

import os
import io
import warnings
import requests
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime

warnings.filterwarnings("ignore")

# ── paths ─────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR  = os.path.join(BASE_DIR, "input", "raw")
PROC_DIR = os.path.join(BASE_DIR, "input", "processed")
os.makedirs(RAW_DIR,  exist_ok=True)
os.makedirs(PROC_DIR, exist_ok=True)

# ── source URLs ───────────────────────────────────────────────────────────────
EFV_BASE = "https://www.efv.admin.ch/dam/en/sd-web"
GMBF_URL        = f"{EFV_BASE}/d9FElh8dV0db/resultate-gmbf.xlsx"
BONDS_URL       = f"{EFV_BASE}/AFFQnqhWiEVW/resultate-anleihen.xlsx"
OUTSTANDING_URL = f"{EFV_BASE}/K9gjjBZzAoPz/ausstehende-anleihen.xlsx"

GMBF_RAW        = os.path.join(RAW_DIR,  "efv_gmbf.xlsx")
BONDS_RAW       = os.path.join(RAW_DIR,  "efv_bonds.xlsx")
OUTSTANDING_RAW = os.path.join(RAW_DIR,  "efv_outstanding.xlsx")

GMBF_PROC        = os.path.join(PROC_DIR, "efv_gmbf_auctions.csv")
BONDS_PROC       = os.path.join(PROC_DIR, "efv_bond_auctions.csv")
OUTSTANDING_PROC = os.path.join(PROC_DIR, "efv_bonds_outstanding.csv")


# ─────────────────────────────────────────────────────────────────────────────
# 1.  DOWNLOAD HELPER
# ─────────────────────────────────────────────────────────────────────────────

def download_file(url: str, dest: str, force: bool = False) -> bytes:
    """Download *url* to *dest*; use cache if available unless force=True."""
    if not force and os.path.exists(dest):
        print(f"  [cache] {os.path.basename(dest)}")
        with open(dest, "rb") as f:
            return f.read()
    print(f"  [download] {url}")
    resp = requests.get(url, verify=False, timeout=120)
    resp.raise_for_status()
    with open(dest, "wb") as f:
        f.write(resp.content)
    print(f"  -> {os.path.basename(dest)}  ({len(resp.content):,} bytes)")
    return resp.content


# ─────────────────────────────────────────────────────────────────────────────
# 2.  GMBF PARSER
# ─────────────────────────────────────────────────────────────────────────────
#
# One sheet per calendar year (2012–present).
# Header block: rows 1–7 (multilingual titles), rows 8–11 (column headers in
#   DE / FR / IT / EN), data from row 12 onward.
#
# Column layout:
#   2012–2025 (11 columns, with Series):
#     A  auction_date
#     B  settlement_date
#     C  maturity_date
#     D  term_days
#     E  series        ← numeric series number; dropped from 2026 onward
#     F  isin
#     G  total_bids_chf_mn
#     H  bids_no_price_chf_mn
#     I  issue_volume_chf_mn
#     J  price
#     K  yield_pa      ← stored as decimal (e.g. 0.003 = 0.3 % p.a.)
#
#   2026+ (10 columns, no Series):
#     A  auction_date
#     B  settlement_date
#     C  maturity_date
#     D  term_days
#     E  isin
#     F  total_bids_chf_mn
#     G  bids_no_price_chf_mn
#     H  issue_volume_chf_mn
#     I  price
#     J  yield_pa
#
# Term buckets (from term_days):
#   ≤ 100 d → "91d"   (3-month)
#   ≤ 200 d → "182d"  (6-month)
#   else    → "364d"  (12-month)

_GMBF_COLS_WITH_SERIES = [
    "auction_date", "settlement_date", "maturity_date", "term_days",
    "series", "isin",
    "total_bids_chf_mn", "bids_no_price_chf_mn", "issue_volume_chf_mn",
    "price", "yield_pa",
]

_GMBF_COLS_NO_SERIES = [
    "auction_date", "settlement_date", "maturity_date", "term_days",
    "isin",
    "total_bids_chf_mn", "bids_no_price_chf_mn", "issue_volume_chf_mn",
    "price", "yield_pa",
]

def _gmbf_term_bucket(days: int) -> str:
    if days <= 100:  return "91d"
    if days <= 200:  return "182d"
    return "364d"

def parse_gmbf(xlsx_bytes: bytes) -> pd.DataFrame:
    """Parse all annual sheets of the GMBF results workbook."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    frames = []

    for sheet_name in wb.sheetnames:
        year = int(sheet_name)
        ws   = wb[sheet_name]
        has_series = (year <= 2025)
        col_names  = _GMBF_COLS_WITH_SERIES if has_series else _GMBF_COLS_NO_SERIES
        n_cols     = len(col_names)

        rows = []
        for row in ws.iter_rows(min_row=12, values_only=True):
            vals = list(row[:n_cols])
            # Skip blank / summary rows (first cell must be a date)
            if not vals or not isinstance(vals[0], datetime):
                continue
            rows.append(vals)

        if not rows:
            continue

        df = pd.DataFrame(rows, columns=col_names)

        # If no series column, insert NaN placeholder for uniformity
        if not has_series:
            df.insert(4, "series", np.nan)

        frames.append(df)

    if not frames:
        return pd.DataFrame()

    df_all = pd.concat(frames, ignore_index=True)

    # Coerce types
    for col in ("auction_date", "settlement_date", "maturity_date"):
        df_all[col] = pd.to_datetime(df_all[col])
    df_all["term_days"] = pd.to_numeric(df_all["term_days"], errors="coerce").astype("Int64")
    for col in ("total_bids_chf_mn", "bids_no_price_chf_mn",
                "issue_volume_chf_mn", "price", "yield_pa"):
        df_all[col] = pd.to_numeric(df_all[col], errors="coerce")

    # Convert yield from decimal to per cent
    df_all["yield_pct"] = (df_all["yield_pa"] * 100).round(6)

    # Bid-to-cover ratio
    df_all["bid_cover"] = (
        df_all["total_bids_chf_mn"] / df_all["issue_volume_chf_mn"]
    ).round(4)

    # Term bucket
    df_all["term_bucket"] = df_all["term_days"].apply(
        lambda d: _gmbf_term_bucket(int(d)) if pd.notna(d) else np.nan
    )

    # Sort chronologically
    df_all = df_all.sort_values("auction_date").reset_index(drop=True)

    # Final column order
    cols = [
        "auction_date", "settlement_date", "maturity_date",
        "term_days", "term_bucket", "series", "isin",
        "total_bids_chf_mn", "bids_no_price_chf_mn", "issue_volume_chf_mn",
        "bid_cover", "price", "yield_pct",
    ]
    return df_all[cols]


# ─────────────────────────────────────────────────────────────────────────────
# 3.  CONFEDERATION BOND PARSER
# ─────────────────────────────────────────────────────────────────────────────
#
# One sheet per year (2011–present), two tranche rows per auction.
# Header block identical structure: rows 8–11 headers, data from row 12.
#
# Column layout (15 cols, last 2 always None):
#   A  auction_date
#   B  settlement_date
#   C  maturity_date
#   D  prov_isin          (provisional ISIN; '-' for new series from 2025 onward
#                          occasionally; otherwise CH…)
#   E  fungible_isin      (existing ISIN this tranche is fungible with)
#   F  bond_name          (e.g. "Eidg. 23.06.21/35")
#   G  coupon_rate        (decimal, e.g. 0.0025 = 0.25 %)
#   H  total_bids_chf_mn
#   I  bids_no_price_chf_mn
#   J  issue_volume_chf_mn
#   K  price
#   L  yield_pa           (decimal)
#   M  own_holdings_not_placed_chf_mn

_BOND_COLS = [
    "auction_date", "settlement_date", "maturity_date",
    "prov_isin", "fungible_isin", "bond_name",
    "coupon_rate",
    "total_bids_chf_mn", "bids_no_price_chf_mn", "issue_volume_chf_mn",
    "price", "yield_pa",
    "own_holdings_not_placed_chf_mn",
]

def parse_bonds(xlsx_bytes: bytes) -> pd.DataFrame:
    """Parse all annual sheets of the Confederation bond auction results workbook."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    frames = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=12, values_only=True):
            vals = list(row[:13])
            if not vals or not isinstance(vals[0], datetime):
                continue
            rows.append(vals)
        if not rows:
            continue
        frames.append(pd.DataFrame(rows, columns=_BOND_COLS))

    if not frames:
        return pd.DataFrame()

    df_all = pd.concat(frames, ignore_index=True)

    # Coerce types
    for col in ("auction_date", "settlement_date", "maturity_date"):
        df_all[col] = pd.to_datetime(df_all[col])
    for col in ("coupon_rate", "total_bids_chf_mn", "bids_no_price_chf_mn",
                "issue_volume_chf_mn", "price", "yield_pa",
                "own_holdings_not_placed_chf_mn"):
        df_all[col] = pd.to_numeric(df_all[col], errors="coerce")

    # Convert rates/yields to per cent
    df_all["coupon_pct"]  = (df_all["coupon_rate"] * 100).round(4)
    df_all["yield_pct"]   = (df_all["yield_pa"]    * 100).round(6)

    # Bid-to-cover
    df_all["bid_cover"] = (
        df_all["total_bids_chf_mn"] / df_all["issue_volume_chf_mn"]
    ).round(4)

    # Residual maturity at auction date (years)
    df_all["residual_maturity_yrs"] = (
        (df_all["maturity_date"] - df_all["auction_date"]).dt.days / 365.25
    ).round(2)

    # Sort chronologically
    df_all = df_all.sort_values(["auction_date", "maturity_date"]).reset_index(drop=True)

    cols = [
        "auction_date", "settlement_date", "maturity_date",
        "residual_maturity_yrs",
        "prov_isin", "fungible_isin", "bond_name",
        "coupon_pct",
        "total_bids_chf_mn", "bids_no_price_chf_mn", "issue_volume_chf_mn",
        "bid_cover", "price", "yield_pct",
        "own_holdings_not_placed_chf_mn",
    ]
    return df_all[cols]


# ─────────────────────────────────────────────────────────────────────────────
# 4.  OUTSTANDING BONDS PARSER
# ─────────────────────────────────────────────────────────────────────────────
#
# Single sheet 'Ausstehende Anleihen'; 8 data columns; data rows 12–34
# (last meaningful row before totals at row 35).

_OUTSTANDING_COLS = [
    "isin", "bond_name", "maturity_date", "coupon_rate",
    "total_issued_incl_own_chf_mn",
    "total_placed_on_market_chf_mn",
    "own_holdings_placed_chf_mn",
    "own_holdings_available_chf_mn",
]

def parse_outstanding(xlsx_bytes: bytes) -> pd.DataFrame:
    """Parse the outstanding bonds snapshot workbook."""
    wb  = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws  = wb.sheetnames[0]
    ws  = wb[ws]

    # Find reference date from the row that starts with 'Stand per'
    ref_date_str = None
    rows_data = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row[:8])
        # Detect reference-date row
        if isinstance(vals[0], str) and vals[0].startswith("Stand"):
            ref_date_str = vals[0]
            continue
        # Skip totals rows (ISIN would be None, value would be numeric)
        if vals[0] is None:
            continue
        # Skip if not a valid ISIN-like string
        if not isinstance(vals[0], str) or not vals[0].startswith("CH"):
            continue
        rows_data.append(vals)

    df = pd.DataFrame(rows_data, columns=_OUTSTANDING_COLS)
    df["maturity_date"]  = pd.to_datetime(df["maturity_date"])
    df["coupon_rate"]    = pd.to_numeric(df["coupon_rate"],    errors="coerce")
    df["coupon_pct"]     = (df["coupon_rate"] * 100).round(4)
    for col in ("total_issued_incl_own_chf_mn", "total_placed_on_market_chf_mn",
                "own_holdings_placed_chf_mn", "own_holdings_available_chf_mn"):
        df[col] = pd.to_numeric(df[col], errors="coerce")

    if ref_date_str:
        # Extract date from e.g. "Stand per/ Etat au/ … situation at: 31.03.2026"
        import re
        m = re.search(r"(\d{2}\.\d{2}\.\d{4})", ref_date_str)
        if m:
            df["snapshot_date"] = pd.to_datetime(m.group(1), dayfirst=True)

    cols = [
        "isin", "bond_name", "maturity_date", "coupon_pct",
        "total_issued_incl_own_chf_mn", "total_placed_on_market_chf_mn",
        "own_holdings_placed_chf_mn", "own_holdings_available_chf_mn",
        "snapshot_date",
    ]
    return df[[c for c in cols if c in df.columns]]


# ─────────────────────────────────────────────────────────────────────────────
# 5.  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("EFV Debt Issuance Analysis")
    print("=" * 60)

    # ── 5a. GMBF ──────────────────────────────────────────────────────────────
    print("\n[1] GMBF auction results")
    gmbf_bytes = download_file(GMBF_URL, GMBF_RAW)
    df_gmbf = parse_gmbf(gmbf_bytes)
    print(f"    Parsed {len(df_gmbf):,} auction rows")
    print(f"    Date range: {df_gmbf['auction_date'].min().date()} "
          f"to {df_gmbf['auction_date'].max().date()}")
    print(f"    Term buckets: {df_gmbf['term_bucket'].value_counts().to_dict()}")
    print(f"    Yield range: {df_gmbf['yield_pct'].min():.4f}% "
          f"to {df_gmbf['yield_pct'].max():.4f}%")
    df_gmbf.to_csv(GMBF_PROC, index=False)
    print(f"    Saved -> {GMBF_PROC}")

    print("\n    Recent auctions (last 8 rows):")
    display_cols = ["auction_date", "term_bucket", "isin",
                    "issue_volume_chf_mn", "bid_cover", "price", "yield_pct"]
    print(df_gmbf[display_cols].tail(8).to_string(index=False))

    # ── 5b. Bonds ─────────────────────────────────────────────────────────────
    print("\n[2] Confederation bond auction results")
    bonds_bytes = download_file(BONDS_URL, BONDS_RAW)
    df_bonds = parse_bonds(bonds_bytes)
    print(f"    Parsed {len(df_bonds):,} tranche rows "
          f"({len(df_bonds['auction_date'].unique())} auction dates)")
    print(f"    Date range: {df_bonds['auction_date'].min().date()} "
          f"to {df_bonds['auction_date'].max().date()}")
    print(f"    Maturity range: {df_bonds['residual_maturity_yrs'].min():.1f} "
          f"to {df_bonds['residual_maturity_yrs'].max():.1f} years")
    df_bonds.to_csv(BONDS_PROC, index=False)
    print(f"    Saved -> {BONDS_PROC}")

    print("\n    Recent auctions (last 8 tranche rows):")
    display_cols_b = ["auction_date", "bond_name", "residual_maturity_yrs",
                      "coupon_pct", "issue_volume_chf_mn",
                      "bid_cover", "price", "yield_pct"]
    print(df_bonds[display_cols_b].tail(8).to_string(index=False))

    # ── 5c. Outstanding bonds ─────────────────────────────────────────────────
    print("\n[3] Outstanding bonds snapshot")
    out_bytes = download_file(OUTSTANDING_URL, OUTSTANDING_RAW)
    df_out = parse_outstanding(out_bytes)
    snap = df_out["snapshot_date"].iloc[0].date() if "snapshot_date" in df_out.columns else "unknown"
    print(f"    {len(df_out)} bonds outstanding as of {snap}")
    total_placed = df_out["total_placed_on_market_chf_mn"].sum()
    total_own    = df_out["own_holdings_available_chf_mn"].sum()
    print(f"    Total placed on market: CHF {total_placed:,.0f} mn")
    print(f"    Total own holdings available: CHF {total_own:,.0f} mn")
    df_out.to_csv(OUTSTANDING_PROC, index=False)
    print(f"    Saved -> {OUTSTANDING_PROC}")

    # ── 5d. Summary ───────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("Done. Output files:")
    for f in [GMBF_PROC, BONDS_PROC, OUTSTANDING_PROC]:
        print(f"  {f}")
    print("=" * 60)

    return df_gmbf, df_bonds, df_out


if __name__ == "__main__":
    df_gmbf, df_bonds, df_out = main()
